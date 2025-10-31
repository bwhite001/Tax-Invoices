"""
Catalog Loader - Load existing invoice catalogs

This module loads previously exported invoice catalogs from CSV, Excel, or JSON.
Follows Single Responsibility Principle.
"""
from pathlib import Path
from typing import List, Dict, Any, Optional
import csv
import json
from datetime import datetime

# Import from parent modules
import sys
sys.path.append(str(Path(__file__).parent.parent))

from utils import get_logger

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class CatalogLoader:
    """
    Load invoice catalogs from various formats
    
    Responsibilities:
    - Load catalog from CSV
    - Load catalog from Excel
    - Load catalog from JSON
    - Validate catalog structure
    
    Does NOT:
    - Modify catalog data
    - Calculate tax deductions
    - Process new invoices
    """
    
    def __init__(self):
        """Initialize catalog loader"""
        self.logger = get_logger()
    
    def load_from_csv(self, csv_path: Path) -> List[Dict[str, Any]]:
        """
        Load catalog from CSV file
        
        Args:
            csv_path: Path to CSV file
        
        Returns:
            List of catalog entries
        """
        csv_path = Path(csv_path)
        
        if not csv_path.exists():
            self.logger.error(f"CSV file not found: {csv_path}")
            return []
        
        self.logger.info(f"Loading catalog from CSV: {csv_path}")
        
        catalog_entries = []
        
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                
                for row in reader:
                    # Convert string values to appropriate types
                    entry = self._parse_csv_row(row)
                    catalog_entries.append(entry)
            
            self.logger.success(f"Loaded {len(catalog_entries)} entries from CSV")
            return catalog_entries
            
        except Exception as e:
            self.logger.error(f"Error loading CSV: {e}")
            return []
    
    def _parse_csv_row(self, row: Dict[str, str]) -> Dict[str, Any]:
        """
        Parse CSV row and convert types
        
        Args:
            row: CSV row as dictionary
        
        Returns:
            Parsed entry with correct types
        """
        entry = {}
        
        # String fields
        string_fields = [
            'FileName', 'FileType', 'FilePath', 'OriginalPath',
            'VendorName', 'VendorABN', 'InvoiceNumber', 'InvoiceDate',
            'DueDate', 'Currency', 'Category', 'ProcessingStatus',
            'FileHash', 'MovedTo'
        ]
        
        for field in string_fields:
            entry[field] = row.get(field, '')
        
        # Numeric fields
        try:
            entry['SubTotal'] = float(row.get('SubTotal', 0))
        except (ValueError, TypeError):
            entry['SubTotal'] = 0.0
        
        try:
            entry['Tax'] = float(row.get('Tax', 0))
        except (ValueError, TypeError):
            entry['Tax'] = 0.0
        
        try:
            entry['TotalAmount'] = float(row.get('TotalAmount', 0))
        except (ValueError, TypeError):
            entry['TotalAmount'] = 0.0
        
        # Boolean fields
        entry['NeedsManualReview'] = row.get('NeedsManualReview', 'False').lower() in ['true', '1', 'yes']
        
        # List fields
        missing_fields = row.get('MissingFields', '')
        if missing_fields:
            entry['MissingFields'] = [f.strip() for f in missing_fields.split(',')]
        else:
            entry['MissingFields'] = []
        
        # DateTime fields
        try:
            processed_dt = row.get('ProcessedDateTime', '')
            if processed_dt:
                entry['ProcessedDateTime'] = datetime.strptime(processed_dt, '%Y-%m-%d %H:%M:%S')
            else:
                entry['ProcessedDateTime'] = None
        except (ValueError, TypeError):
            entry['ProcessedDateTime'] = None
        
        return entry
    
    def load_from_excel(self, excel_path: Path, sheet_name: str = "Catalog") -> List[Dict[str, Any]]:
        """
        Load catalog from Excel file
        
        Args:
            excel_path: Path to Excel file
            sheet_name: Name of sheet to load (default: "Catalog")
        
        Returns:
            List of catalog entries
        """
        if not EXCEL_AVAILABLE:
            self.logger.error("openpyxl not available, cannot load Excel files")
            return []
        
        excel_path = Path(excel_path)
        
        if not excel_path.exists():
            self.logger.error(f"Excel file not found: {excel_path}")
            return []
        
        self.logger.info(f"Loading catalog from Excel: {excel_path} (sheet: {sheet_name})")
        
        catalog_entries = []
        
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            
            if sheet_name not in wb.sheetnames:
                self.logger.error(f"Sheet '{sheet_name}' not found in Excel file")
                return []
            
            ws = wb[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                entry = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        header = headers[i]
                        entry[self._normalize_header(header)] = value
                
                # Parse entry
                parsed_entry = self._parse_excel_row(entry)
                catalog_entries.append(parsed_entry)
            
            wb.close()
            
            self.logger.success(f"Loaded {len(catalog_entries)} entries from Excel")
            return catalog_entries
            
        except Exception as e:
            self.logger.error(f"Error loading Excel: {e}")
            return []
    
    def _normalize_header(self, header: str) -> str:
        """
        Normalize Excel header to match CSV field names
        
        Args:
            header: Excel header
        
        Returns:
            Normalized field name
        """
        # Map Excel headers to field names
        header_map = {
            'File Name': 'FileName',
            'Vendor Name': 'VendorName',
            'Invoice Number': 'InvoiceNumber',
            'Invoice Date': 'InvoiceDate',
            'Due Date': 'DueDate',
            'SubTotal': 'SubTotal',
            'Tax': 'Tax',
            'Total Amount': 'TotalAmount',
            'Currency': 'Currency',
            'Category': 'Category',
            'Status': 'ProcessingStatus',
            'Needs Review': 'NeedsManualReview',
            'Missing Fields': 'MissingFields',
            'File Path': 'FilePath'
        }
        
        return header_map.get(header, header)
    
    def _parse_excel_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """
        Parse Excel row and convert types
        
        Args:
            row: Excel row as dictionary
        
        Returns:
            Parsed entry with correct types
        """
        entry = {}
        
        # String fields
        string_fields = [
            'FileName', 'FileType', 'FilePath', 'OriginalPath',
            'VendorName', 'VendorABN', 'InvoiceNumber', 'InvoiceDate',
            'DueDate', 'Currency', 'Category', 'ProcessingStatus',
            'FileHash', 'MovedTo'
        ]
        
        for field in string_fields:
            value = row.get(field, '')
            entry[field] = str(value) if value is not None else ''
        
        # Numeric fields
        entry['SubTotal'] = float(row.get('SubTotal', 0)) if row.get('SubTotal') else 0.0
        entry['Tax'] = float(row.get('Tax', 0)) if row.get('Tax') else 0.0
        entry['TotalAmount'] = float(row.get('TotalAmount', 0)) if row.get('TotalAmount') else 0.0
        
        # Boolean fields
        needs_review = row.get('NeedsManualReview', 'No')
        entry['NeedsManualReview'] = str(needs_review).lower() in ['yes', 'true', '1']
        
        # List fields
        missing_fields = row.get('MissingFields', '')
        if missing_fields:
            entry['MissingFields'] = [f.strip() for f in str(missing_fields).split(',')]
        else:
            entry['MissingFields'] = []
        
        # DateTime fields
        entry['ProcessedDateTime'] = row.get('ProcessedDateTime')
        
        return entry
    
    def load_from_json(self, json_path: Path) -> List[Dict[str, Any]]:
        """
        Load catalog from JSON file
        
        Args:
            json_path: Path to JSON file
        
        Returns:
            List of catalog entries
        """
        json_path = Path(json_path)
        
        if not json_path.exists():
            self.logger.error(f"JSON file not found: {json_path}")
            return []
        
        self.logger.info(f"Loading catalog from JSON: {json_path}")
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                catalog_entries = json.load(f)
            
            # Parse datetime strings
            for entry in catalog_entries:
                if 'ProcessedDateTime' in entry and isinstance(entry['ProcessedDateTime'], str):
                    try:
                        entry['ProcessedDateTime'] = datetime.fromisoformat(entry['ProcessedDateTime'])
                    except (ValueError, TypeError):
                        entry['ProcessedDateTime'] = None
            
            self.logger.success(f"Loaded {len(catalog_entries)} entries from JSON")
            return catalog_entries
            
        except Exception as e:
            self.logger.error(f"Error loading JSON: {e}")
            return []
    
    def load_catalog(self, file_path: Path) -> List[Dict[str, Any]]:
        """
        Load catalog from file (auto-detect format)
        
        Args:
            file_path: Path to catalog file
        
        Returns:
            List of catalog entries
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            self.logger.error(f"File not found: {file_path}")
            return []
        
        # Detect format by extension
        ext = file_path.suffix.lower()
        
        if ext == '.csv':
            return self.load_from_csv(file_path)
        elif ext in ['.xlsx', '.xls']:
            return self.load_from_excel(file_path)
        elif ext == '.json':
            return self.load_from_json(file_path)
        else:
            self.logger.error(f"Unsupported file format: {ext}")
            return []
    
    def validate_catalog(self, catalog_entries: List[Dict[str, Any]]) -> tuple[bool, List[str]]:
        """
        Validate catalog structure and data
        
        Args:
            catalog_entries: List of catalog entries
        
        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        errors = []
        
        if not catalog_entries:
            errors.append("Catalog is empty")
            return False, errors
        
        # Required fields
        required_fields = [
            'FileName', 'VendorName', 'InvoiceDate', 'TotalAmount', 'Category'
        ]
        
        # Check each entry
        for i, entry in enumerate(catalog_entries, 1):
            # Check required fields
            for field in required_fields:
                if field not in entry:
                    errors.append(f"Entry {i}: Missing required field '{field}'")
            
            # Validate data types
            if 'TotalAmount' in entry:
                try:
                    float(entry['TotalAmount'])
                except (ValueError, TypeError):
                    errors.append(f"Entry {i}: Invalid TotalAmount value")
        
        is_valid = len(errors) == 0
        
        if is_valid:
            self.logger.success("Catalog validation passed")
        else:
            self.logger.warning(f"Catalog validation failed with {len(errors)} error(s)")
        
        return is_valid, errors
    
    def get_catalog_summary(self, catalog_entries: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Get summary statistics for catalog
        
        Args:
            catalog_entries: List of catalog entries
        
        Returns:
            Dictionary with summary statistics
        """
        if not catalog_entries:
            return {
                'total_entries': 0,
                'total_amount': 0.0,
                'categories': {},
                'date_range': None
            }
        
        # Calculate statistics
        total_amount = sum(entry.get('TotalAmount', 0) for entry in catalog_entries)
        
        # Count by category
        categories = {}
        for entry in catalog_entries:
            category = entry.get('Category', 'Unknown')
            if category not in categories:
                categories[category] = {'count': 0, 'total': 0.0}
            categories[category]['count'] += 1
            categories[category]['total'] += entry.get('TotalAmount', 0)
        
        # Get date range
        dates = [entry.get('InvoiceDate', '') for entry in catalog_entries if entry.get('InvoiceDate')]
        date_range = None
        if dates:
            dates.sort()
            date_range = {'start': dates[0], 'end': dates[-1]}
        
        summary = {
            'total_entries': len(catalog_entries),
            'total_amount': total_amount,
            'categories': categories,
            'date_range': date_range
        }
        
        return summary
