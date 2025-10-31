"""
Catalog Exporter - Export invoice catalog WITHOUT tax calculations

This module exports the invoice catalog in various formats (CSV, Excel)
WITHOUT any tax-related fields. Follows Single Responsibility Principle.
"""
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
import csv
import json

# Import from parent modules
import sys
sys.path.append(str(Path(__file__).parent.parent))

from utils import get_logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class CatalogExporter:
    """
    Export invoice catalog without tax calculations
    
    Responsibilities:
    - Export catalog to CSV
    - Export catalog to Excel
    - Export manual review list
    
    Does NOT:
    - Include tax deduction fields
    - Include ATO references
    - Include claim methods
    """
    
    def __init__(self, output_folder: Path):
        """
        Initialize exporter
        
        Args:
            output_folder: Output folder path
        """
        self.output_folder = Path(output_folder)
        self.logger = get_logger()
        
        # Ensure output folder exists
        self.output_folder.mkdir(parents=True, exist_ok=True)
    
    def export_csv(self, catalog_entries: List[Dict[str, Any]]) -> tuple[Path, Path, Path]:
        """
        Export catalog to CSV files
        
        Args:
            catalog_entries: List of catalog entries
        
        Returns:
            Tuple of (catalog_path, summary_path, manual_review_path)
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Export main catalog
        catalog_path = self._export_catalog_csv(catalog_entries, timestamp)
        
        # Export summary
        summary_path = self._export_summary_csv(catalog_entries, timestamp)
        
        # Export manual review list
        manual_review_path = self._export_manual_review_csv(catalog_entries, timestamp)
        
        return catalog_path, summary_path, manual_review_path
    
    def _export_catalog_csv(self, catalog_entries: List[Dict[str, Any]], 
                           timestamp: str) -> Path:
        """
        Export main catalog CSV (NO tax fields)
        
        Args:
            catalog_entries: List of catalog entries
            timestamp: Timestamp string
        
        Returns:
            Path to exported CSV file
        """
        catalog_path = self.output_folder / f"Invoice_Catalog_{timestamp}.csv"
        
        # Define catalog fields (NO tax fields)
        fieldnames = [
            'FileName',
            'FileType',
            'FilePath',
            'OriginalPath',
            'ProcessedDateTime',
            'VendorName',
            'VendorABN',
            'InvoiceNumber',
            'InvoiceDate',
            'DueDate',
            'SubTotal',
            'Tax',
            'TotalAmount',
            'Currency',
            'Category',
            'ProcessingStatus',
            'FileHash',
            'MovedTo',
            'NeedsManualReview',
            'MissingFields'
        ]
        
        with open(catalog_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for entry in catalog_entries:
                # Convert lists to strings
                row = entry.copy()
                if isinstance(row.get('MissingFields'), list):
                    row['MissingFields'] = ', '.join(row['MissingFields'])
                
                # Format datetime
                if isinstance(row.get('ProcessedDateTime'), datetime):
                    row['ProcessedDateTime'] = row['ProcessedDateTime'].strftime('%Y-%m-%d %H:%M:%S')
                
                writer.writerow(row)
        
        self.logger.success(f"Catalog CSV exported: {catalog_path}")
        return catalog_path
    
    def _export_summary_csv(self, catalog_entries: List[Dict[str, Any]], 
                           timestamp: str) -> Path:
        """
        Export summary CSV with statistics
        
        Args:
            catalog_entries: List of catalog entries
            timestamp: Timestamp string
        
        Returns:
            Path to exported CSV file
        """
        summary_path = self.output_folder / f"Catalog_Summary_{timestamp}.csv"
        
        # Calculate statistics
        total_invoices = len(catalog_entries)
        total_amount = sum(entry.get('TotalAmount', 0) for entry in catalog_entries)
        
        # Count by category
        category_stats = {}
        for entry in catalog_entries:
            category = entry.get('Category', 'Unknown')
            if category not in category_stats:
                category_stats[category] = {'count': 0, 'total': 0.0}
            category_stats[category]['count'] += 1
            category_stats[category]['total'] += entry.get('TotalAmount', 0)
        
        # Count by status
        status_stats = {}
        for entry in catalog_entries:
            status = entry.get('ProcessingStatus', 'Unknown')
            status_stats[status] = status_stats.get(status, 0) + 1
        
        # Write summary
        with open(summary_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Overall statistics
            writer.writerow(['INVOICE CATALOG SUMMARY'])
            writer.writerow([])
            writer.writerow(['Total Invoices', total_invoices])
            writer.writerow(['Total Amount', f'${total_amount:,.2f}'])
            writer.writerow([])
            
            # Category breakdown
            writer.writerow(['CATEGORY BREAKDOWN'])
            writer.writerow(['Category', 'Count', 'Total Amount'])
            for category, stats in sorted(category_stats.items()):
                writer.writerow([
                    category,
                    stats['count'],
                    f"${stats['total']:,.2f}"
                ])
            writer.writerow([])
            
            # Status breakdown
            writer.writerow(['PROCESSING STATUS'])
            writer.writerow(['Status', 'Count'])
            for status, count in sorted(status_stats.items()):
                writer.writerow([status, count])
        
        self.logger.success(f"Summary CSV exported: {summary_path}")
        return summary_path
    
    def _export_manual_review_csv(self, catalog_entries: List[Dict[str, Any]], 
                                  timestamp: str) -> Path:
        """
        Export manual review list CSV
        
        Args:
            catalog_entries: List of catalog entries
            timestamp: Timestamp string
        
        Returns:
            Path to exported CSV file
        """
        manual_review_path = self.output_folder / f"Manual_Review_Required_{timestamp}.csv"
        
        # Filter entries needing manual review
        review_entries = [
            entry for entry in catalog_entries 
            if entry.get('NeedsManualReview', False)
        ]
        
        if not review_entries:
            self.logger.info("No entries require manual review")
            return manual_review_path
        
        # Define fields for manual review
        fieldnames = [
            'FileName',
            'VendorName',
            'InvoiceDate',
            'TotalAmount',
            'Category',
            'MissingFields',
            'FilePath'
        ]
        
        with open(manual_review_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for entry in review_entries:
                row = {
                    'FileName': entry.get('FileName', ''),
                    'VendorName': entry.get('VendorName', ''),
                    'InvoiceDate': entry.get('InvoiceDate', ''),
                    'TotalAmount': entry.get('TotalAmount', 0),
                    'Category': entry.get('Category', ''),
                    'MissingFields': ', '.join(entry.get('MissingFields', [])) if isinstance(entry.get('MissingFields'), list) else entry.get('MissingFields', ''),
                    'FilePath': entry.get('FilePath', '')
                }
                writer.writerow(row)
        
        self.logger.success(f"Manual review CSV exported: {manual_review_path} ({len(review_entries)} entries)")
        return manual_review_path
    
    def export_excel(self, catalog_entries: List[Dict[str, Any]]) -> Path:
        """
        Export catalog to Excel file
        
        Args:
            catalog_entries: List of catalog entries
        
        Returns:
            Path to exported Excel file
        """
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl not available, skipping Excel export")
            return None
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = self.output_folder / f"Invoice_Catalog_{timestamp}.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create catalog sheet
        self._create_catalog_sheet(wb, catalog_entries)
        
        # Create summary sheet
        self._create_summary_sheet(wb, catalog_entries)
        
        # Create manual review sheet
        self._create_manual_review_sheet(wb, catalog_entries)
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Save workbook
        wb.save(excel_path)
        
        self.logger.success(f"Excel file exported: {excel_path}")
        return excel_path
    
    def _create_catalog_sheet(self, wb, catalog_entries: List[Dict[str, Any]]):
        """Create catalog sheet in Excel workbook"""
        ws = wb.create_sheet("Catalog", 0)
        
        # Define headers (NO tax fields)
        headers = [
            'File Name', 'Vendor Name', 'Invoice Number', 'Invoice Date',
            'Due Date', 'SubTotal', 'Tax', 'Total Amount', 'Currency',
            'Category', 'Status', 'Needs Review', 'Missing Fields', 'File Path'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, entry in enumerate(catalog_entries, 2):
            ws.cell(row=row_idx, column=1, value=entry.get('FileName', ''))
            ws.cell(row=row_idx, column=2, value=entry.get('VendorName', ''))
            ws.cell(row=row_idx, column=3, value=entry.get('InvoiceNumber', ''))
            ws.cell(row=row_idx, column=4, value=entry.get('InvoiceDate', ''))
            ws.cell(row=row_idx, column=5, value=entry.get('DueDate', ''))
            ws.cell(row=row_idx, column=6, value=entry.get('SubTotal', 0))
            ws.cell(row=row_idx, column=7, value=entry.get('Tax', 0))
            ws.cell(row=row_idx, column=8, value=entry.get('TotalAmount', 0))
            ws.cell(row=row_idx, column=9, value=entry.get('Currency', 'AUD'))
            ws.cell(row=row_idx, column=10, value=entry.get('Category', ''))
            ws.cell(row=row_idx, column=11, value=entry.get('ProcessingStatus', ''))
            ws.cell(row=row_idx, column=12, value='Yes' if entry.get('NeedsManualReview', False) else 'No')
            
            missing_fields = entry.get('MissingFields', [])
            if isinstance(missing_fields, list):
                missing_fields = ', '.join(missing_fields)
            ws.cell(row=row_idx, column=13, value=missing_fields)
            ws.cell(row=row_idx, column=14, value=entry.get('FilePath', ''))
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, wb, catalog_entries: List[Dict[str, Any]]):
        """Create summary sheet in Excel workbook"""
        ws = wb.create_sheet("Summary", 1)
        
        # Calculate statistics
        total_invoices = len(catalog_entries)
        total_amount = sum(entry.get('TotalAmount', 0) for entry in catalog_entries)
        
        # Count by category
        category_stats = {}
        for entry in catalog_entries:
            category = entry.get('Category', 'Unknown')
            if category not in category_stats:
                category_stats[category] = {'count': 0, 'total': 0.0}
            category_stats[category]['count'] += 1
            category_stats[category]['total'] += entry.get('TotalAmount', 0)
        
        # Write summary
        row = 1
        ws.cell(row=row, column=1, value="INVOICE CATALOG SUMMARY").font = Font(bold=True, size=14)
        row += 2
        
        ws.cell(row=row, column=1, value="Total Invoices:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_invoices)
        row += 1
        
        ws.cell(row=row, column=1, value="Total Amount:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"${total_amount:,.2f}")
        row += 2
        
        ws.cell(row=row, column=1, value="CATEGORY BREAKDOWN").font = Font(bold=True, size=12)
        row += 1
        
        ws.cell(row=row, column=1, value="Category").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Count").font = Font(bold=True)
        ws.cell(row=row, column=3, value="Total Amount").font = Font(bold=True)
        row += 1
        
        for category, stats in sorted(category_stats.items()):
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=stats['count'])
            ws.cell(row=row, column=3, value=f"${stats['total']:,.2f}")
            row += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_manual_review_sheet(self, wb, catalog_entries: List[Dict[str, Any]]):
        """Create manual review sheet in Excel workbook"""
        ws = wb.create_sheet("Manual Review", 2)
        
        # Filter entries needing manual review
        review_entries = [
            entry for entry in catalog_entries 
            if entry.get('NeedsManualReview', False)
        ]
        
        if not review_entries:
            ws.cell(row=1, column=1, value="No entries require manual review").font = Font(bold=True)
            return
        
        # Write headers
        headers = ['File Name', 'Vendor Name', 'Invoice Date', 'Total Amount', 
                  'Category', 'Missing Fields', 'File Path']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write data
        for row_idx, entry in enumerate(review_entries, 2):
            ws.cell(row=row_idx, column=1, value=entry.get('FileName', ''))
            ws.cell(row=row_idx, column=2, value=entry.get('VendorName', ''))
            ws.cell(row=row_idx, column=3, value=entry.get('InvoiceDate', ''))
            ws.cell(row=row_idx, column=4, value=entry.get('TotalAmount', 0))
            ws.cell(row=row_idx, column=5, value=entry.get('Category', ''))
            
            missing_fields = entry.get('MissingFields', [])
            if isinstance(missing_fields, list):
                missing_fields = ', '.join(missing_fields)
            ws.cell(row=row_idx, column=6, value=missing_fields)
            ws.cell(row=row_idx, column=7, value=entry.get('FilePath', ''))
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_json(self, catalog_entries: List[Dict[str, Any]]) -> Path:
        """
        Export catalog to JSON file
        
        Args:
            catalog_entries: List of catalog entries
        
        Returns:
            Path to exported JSON file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_path = self.output_folder / f"Invoice_Catalog_{timestamp}.json"
        
        # Convert datetime objects to strings
        entries_for_json = []
        for entry in catalog_entries:
            entry_copy = entry.copy()
            if isinstance(entry_copy.get('ProcessedDateTime'), datetime):
                entry_copy['ProcessedDateTime'] = entry_copy['ProcessedDateTime'].isoformat()
            entries_for_json.append(entry_copy)
        
        # Write JSON
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(entries_for_json, f, indent=2, ensure_ascii=False)
        
        self.logger.success(f"JSON file exported: {json_path}")
        return json_path
