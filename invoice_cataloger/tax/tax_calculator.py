"""
Tax Calculator - Main orchestrator for tax deduction calculations

Coordinates:
- Loading catalog data
- Applying tax strategies
- WFH log integration
- Exporting tax reports
"""
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

# Import from parent modules
import sys
sys.path.append(str(Path(__file__).parent.parent))

from catalog import CatalogLoader
from .strategies.base_strategy import TaxStrategy
from .wfh.wfh_parser import WFHParser
from .wfh.wfh_calculator import WFHCalculator
from utils import get_logger


class TaxCalculator:
    """
    Main tax calculator that orchestrates everything
    
    Features:
    - Load catalog from CSV, Excel, or JSON
    - Apply tax strategy to calculate deductions
    - Support dynamic work-use % from WFH log
    - Export tax report with deductions
    """
    
    def __init__(self, strategy: TaxStrategy, 
                 work_use_percentage: Optional[float] = None,
                 wfh_log_path: Optional[Path] = None,
                 financial_year: Optional[str] = None):
        """
        Initialize tax calculator
        
        Args:
            strategy: Tax calculation strategy (e.g., ATOStrategy)
            work_use_percentage: Static work-use percentage (0-100)
            wfh_log_path: Path to WFH log file (optional)
            financial_year: Financial year for filtering WFH log (optional)
        """
        self.strategy = strategy
        self.logger = get_logger()
        self.financial_year = financial_year
        self.wfh_stats = None
        
        # Calculate work-use percentage
        if wfh_log_path:
            self.work_use_percentage = self._calculate_dynamic_work_use(
                wfh_log_path, financial_year
            )
        else:
            self.work_use_percentage = work_use_percentage or 60.0
            self.logger.info(f"Using static work-use percentage: {self.work_use_percentage}%")
    
    def _calculate_dynamic_work_use(self, wfh_log_path: Path, 
                                   financial_year: Optional[str]) -> float:
        """
        Calculate dynamic work-use percentage from WFH log
        
        Args:
            wfh_log_path: Path to WFH log file
            financial_year: Financial year for filtering (optional)
        
        Returns:
            Calculated work-use percentage
        """
        self.logger.info(f"Loading WFH log: {wfh_log_path}")
        
        # Parse WFH log
        parser = WFHParser()
        log_data = parser.parse(wfh_log_path)
        
        if not log_data:
            self.logger.warning("No WFH data found, using default 60%")
            return 60.0
        
        # Validate log
        is_valid, errors = parser.validate_log(log_data)
        if not is_valid:
            self.logger.warning(f"WFH log validation errors: {', '.join(errors)}")
        
        # Filter by financial year if provided
        if financial_year:
            log_data = parser.filter_by_financial_year(log_data, financial_year)
            
            if not log_data:
                self.logger.warning(f"No WFH data for FY{financial_year}, using default 60%")
                return 60.0
        
        # Calculate percentage
        calculator = WFHCalculator()
        percentage = calculator.calculate_wfh_percentage(log_data)
        
        # Store stats for reporting
        self.wfh_stats = calculator.get_summary_stats(log_data)
        
        # Validate percentage
        is_valid, message = calculator.validate_percentage(percentage)
        if not is_valid:
            self.logger.warning(message)
        else:
            self.logger.info(message)
        
        self.logger.success(f"Using dynamic work-use percentage: {percentage}%")
        
        # Log summary
        self.logger.info(f"WFH Days: {self.wfh_stats['wfh_days']} / {self.wfh_stats['total_days']} total")
        
        return percentage
    
    def calculate_deductions(self, catalog_entries: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Calculate tax deductions for all catalog entries
        
        Args:
            catalog_entries: List of invoice entries from catalog
        
        Returns:
            List of entries with tax deduction data added
        """
        if not catalog_entries:
            self.logger.warning("No catalog entries to process")
            return []
        
        self.logger.info(f"Calculating deductions for {len(catalog_entries)} invoices")
        self.logger.info(f"Using strategy: {self.strategy.get_strategy_name()}")
        self.logger.info(f"Work-use percentage: {self.work_use_percentage}%")
        
        tax_entries = []
        
        for entry in catalog_entries:
            try:
                # Get category
                category = entry.get('Category', 'Other')
                
                # Calculate deduction using strategy
                deduction = self.strategy.calculate_deduction(
                    entry,
                    category,
                    self.work_use_percentage
                )
                
                # Merge catalog entry with deduction data
                tax_entry = {**entry, **deduction}
                tax_entries.append(tax_entry)
            
            except Exception as e:
                self.logger.error(f"Error calculating deduction for {entry.get('FileName', 'unknown')}: {e}")
                
                # Add entry with error status
                tax_entry = {
                    **entry,
                    'DeductibleAmount': 0.00,
                    'ClaimMethod': 'Error',
                    'ClaimNotes': f'Calculation error: {str(e)}',
                    'WorkUsePercentage': 0,
                    'AtoReference': 'N/A',
                    'RequiresDocumentation': ['Manual review required']
                }
                tax_entries.append(tax_entry)
        
        # Log summary
        total_deductible = sum(entry.get('DeductibleAmount', 0) for entry in tax_entries)
        self.logger.success(f"Calculated deductions for {len(tax_entries)} invoices")
        self.logger.info(f"Total deductible amount: ${total_deductible:,.2f}")
        
        return tax_entries
    
    def export_tax_report(self, tax_entries: List[Dict[str, Any]], 
                         output_path: Path,
                         format: str = 'csv') -> Path:
        """
        Export tax report with deductions
        
        Args:
            tax_entries: List of entries with tax data
            output_path: Output file path
            format: Export format ('csv', 'excel', 'json')
        
        Returns:
            Path to exported file
        """
        self.logger.info(f"Exporting tax report to: {output_path}")
        
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        if format == 'csv':
            return self._export_csv(tax_entries, output_path)
        elif format == 'excel':
            return self._export_excel(tax_entries, output_path)
        elif format == 'json':
            return self._export_json(tax_entries, output_path)
        else:
            self.logger.error(f"Unsupported export format: {format}")
            return output_path
    
    def _export_csv(self, tax_entries: List[Dict[str, Any]], output_path: Path) -> Path:
        """Export to CSV format"""
        import csv
        
        if not tax_entries:
            self.logger.warning("No entries to export")
            return output_path
        
        # Define CSV columns
        columns = [
            'FileName', 'VendorName', 'InvoiceDate', 'InvoiceNumber',
            'TotalAmount', 'Category', 'WorkUsePercentage', 'DeductibleAmount',
            'ClaimMethod', 'ClaimNotes', 'AtoReference'
        ]
        
        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
                writer.writeheader()
                writer.writerows(tax_entries)
            
            self.logger.success(f"CSV exported: {output_path}")
            return output_path
        
        except Exception as e:
            self.logger.error(f"Error exporting CSV: {e}")
            return output_path
    
    def _export_excel(self, tax_entries: List[Dict[str, Any]], output_path: Path) -> Path:
        """Export to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tax Report"
            
            # Define columns
            columns = [
                'FileName', 'VendorName', 'InvoiceDate', 'InvoiceNumber',
                'TotalAmount', 'Category', 'WorkUsePercentage', 'DeductibleAmount',
                'ClaimMethod', 'ClaimNotes', 'AtoReference'
            ]
            
            # Write header
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Write data
            for row_idx, entry in enumerate(tax_entries, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = entry.get(col_name, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            self.logger.success(f"Excel exported: {output_path}")
            return output_path
        
        except ImportError:
            self.logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return output_path
        except Exception as e:
            self.logger.error(f"Error exporting Excel: {e}")
            return output_path
    
    def _export_json(self, tax_entries: List[Dict[str, Any]], output_path: Path) -> Path:
        """Export to JSON format"""
        import json
        
        try:
            # Convert datetime objects to strings
            def json_serializer(obj):
                if isinstance(obj, datetime):
                    return obj.isoformat()
                raise TypeError(f"Type {type(obj)} not serializable")
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(tax_entries, f, indent=2, default=json_serializer)
            
            self.logger.success(f"JSON exported: {output_path}")
            return output_path
        
        except Exception as e:
            self.logger.error(f"Error exporting JSON: {e}")
            return output_path
    
    def get_wfh_report(self) -> Optional[str]:
        """
        Get WFH statistics report
        
        Returns:
            Formatted WFH report string or None if no WFH data
        """
        if not self.wfh_stats:
            return None
        
        calculator = WFHCalculator()
        
        # Reconstruct log data from stats for report generation
        # (This is a simplified version - in practice, we'd store the full log data)
        report_lines = [
            "=" * 60,
            "WFH STATISTICS SUMMARY",
            "=" * 60,
        ]
        
        if self.financial_year:
            report_lines.append(f"Financial Year: {self.financial_year}")
        
        report_lines.extend([
            "",
            "OVERALL STATISTICS",
            "-" * 60,
            f"Total Work Days:    {self.wfh_stats['total_days']:>6}",
            f"WFH Days:           {self.wfh_stats['wfh_days']:>6}",
            f"Office Days:        {self.wfh_stats['office_days']:>6}",
            f"WFH Percentage:     {self.wfh_stats['percentage']:>5.1f}%",
            "",
            "=" * 60,
        ])
        
        return "\n".join(report_lines)
