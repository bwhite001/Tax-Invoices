"""
Report Generator Module

Single Responsibility: Generate Excel reports.
Dependency Inversion: Depends on abstractions, not concrete implementations.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportGenerator:
    """
    Generate comprehensive Excel reports
    
    Responsibilities:
    - Create Excel workbook with multiple sheets
    - Format data appropriately
    - Apply consistent styling
    """

    def __init__(self, config):
        """
        Initialize report generator
        
        Args:
            config: ReportConfig object with formatting settings
        """
        self.config = config
        self.workbook = None

    def create_workbook(self) -> openpyxl.Workbook:
        """
        Create new Excel workbook
        
        Returns:
            OpenPyXL workbook object
        """
        self.workbook = openpyxl.Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        return self.workbook

    def add_summary_sheet(self, wfh_stats: Dict[str, Any],
                         invoice_stats: Dict[str, Any],
                         bank_stats: Dict[str, Any]) -> None:
        """
        Add summary overview sheet
        
        Args:
            wfh_stats: WFH statistics
            invoice_stats: Invoice statistics
            bank_stats: Bank statement statistics
        """
        ws = self.workbook.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = f"TAX DEDUCTION SUMMARY - {self.config.financial_year.upper()}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # WFH Statistics
        ws[f'A{row}'] = "WORK FROM HOME STATISTICS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        wfh_data = [
            ("Total Work Days:", wfh_stats.get('total_days', 0)),
            ("WFH Days:", wfh_stats.get('wfh_days', 0)),
            ("Office Days:", wfh_stats.get('office_days', 0)),
            ("WFH Percentage:", f"{wfh_stats.get('wfh_percentage', 0):.2f}%")
        ]
        
        for label, value in wfh_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1  # Empty row
        
        # Invoice Statistics
        ws[f'A{row}'] = "INVOICE STATISTICS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        invoice_data = [
            ("Total Invoiced:", f"${invoice_stats.get('total_invoiced', 0):,.2f}"),
            ("Original Deductions:", f"${invoice_stats.get('original_deduction', 0):,.2f}"),
            ("Recalculated Deductions:", f"${invoice_stats.get('recalculated_deduction', 0):,.2f}"),
            ("Adjustment:", f"${invoice_stats.get('adjustment', 0):,.2f}"),
            ("Deduction Rate:", f"{invoice_stats.get('deduction_rate', 0):.1f}%")
        ]
        
        for label, value in invoice_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1  # Empty row
        
        # Bank Statistics (if available)
        if bank_stats.get('has_data', False):
            ws[f'A{row}'] = "BANK STATEMENT STATISTICS"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            bank_data = [
                ("Total Transactions:", bank_stats.get('total_transactions', 0)),
                ("Tax Relevant:", bank_stats.get('tax_relevant_count', 0)),
                ("Total Amount:", f"${bank_stats.get('total_amount', 0):,.2f}")
            ]
            
            for label, value in bank_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
        
        # Auto-size columns
        self._auto_size_columns(ws, ['A', 'B'])

    def add_invoice_sheet(self, invoice_df: pd.DataFrame) -> None:
        """
        Add detailed invoice catalog sheet
        
        Args:
            invoice_df: Invoice DataFrame
        """
        ws = self.workbook.create_sheet("Invoice Catalog")
        
        # Write dataframe to sheet
        for r in dataframe_to_rows(invoice_df, index=False, header=True):
            ws.append(r)
        
        # Format header
        self._format_header_row(ws)
        
        # Auto-size columns
        self._auto_size_columns(ws)

    def add_category_breakdown_sheet(self, category_df: pd.DataFrame) -> None:
        """
        Add category breakdown sheet
        
        Args:
            category_df: Category summary DataFrame
        """
        ws = self.workbook.create_sheet("Category Breakdown")
        
        # Write dataframe to sheet
        for r in dataframe_to_rows(category_df, index=False, header=True):
            ws.append(r)
        
        # Format header
        self._format_header_row(ws)
        
        # Format currency columns
        currency_cols = ['C', 'D', 'E', 'F']  # Total_Invoiced, Original_Deduction, Recalculated_Deduction, Adjustment
        for col in currency_cols:
            for cell in ws[col]:
                if cell.row > 1:  # Skip header
                    cell.number_format = self.config.currency_format
        
        # Auto-size columns
        self._auto_size_columns(ws)

    def add_wfh_analysis_sheet(self, wfh_data: List[Dict[str, Any]],
                              monthly_stats: Dict[str, Any]) -> None:
        """
        Add WFH analysis sheet
        
        Args:
            wfh_data: WFH daily data
            monthly_stats: Monthly WFH statistics
        """
        ws = self.workbook.create_sheet("WFH Analysis")
        
        # Daily data
        ws['A1'] = "DAILY WFH LOG"
        ws['A1'].font = Font(size=12, bold=True)
        
        # Write daily data
        daily_df = pd.DataFrame(wfh_data)
        if not daily_df.empty:
            daily_df = daily_df[['Date', 'Day', 'Location', 'WorkFromHome']]
            daily_df.columns = ['Date', 'Day', 'Location', 'Work From Home']
            
            row_offset = 3
            for r_idx, r in enumerate(dataframe_to_rows(daily_df, index=False, header=True), row_offset):
                ws.append(r)
            
            # Format daily header
            for cell in ws[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color=self.config.excel_header_color,
                    end_color=self.config.excel_header_color,
                    fill_type="solid"
                )
                cell.font = Font(color=self.config.excel_header_font_color, bold=True)
        
        # Monthly summary
        monthly_start_row = len(wfh_data) + 6
        ws[f'A{monthly_start_row}'] = "MONTHLY SUMMARY"
        ws[f'A{monthly_start_row}'].font = Font(size=12, bold=True)
        
        monthly_start_row += 1
        headers = ['Month', 'Total Days', 'WFH Days', 'Office Days', 'WFH %']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=monthly_start_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=self.config.excel_header_color,
                end_color=self.config.excel_header_color,
                fill_type="solid"
            )
            cell.font = Font(color=self.config.excel_header_font_color, bold=True)
        
        # Write monthly data
        row = monthly_start_row + 1
        for month in sorted(monthly_stats.keys()):
            stats = monthly_stats[month]
            try:
                month_name = datetime.strptime(month + "-01", '%Y-%m-%d').strftime('%B %Y')
            except:
                month_name = month
            
            ws[f'A{row}'] = month_name
            ws[f'B{row}'] = stats['total']
            ws[f'C{row}'] = stats['wfh']
            ws[f'D{row}'] = stats['office']
            ws[f'E{row}'] = f"{stats['percentage']:.1f}%"
            row += 1
        
        # Auto-size columns
        self._auto_size_columns(ws)

    def add_bank_sheet(self, bank_df: pd.DataFrame) -> None:
        """
        Add bank statement sheet
        
        Args:
            bank_df: Bank statement DataFrame
        """
        if bank_df.empty:
            return
        
        ws = self.workbook.create_sheet("Bank Statements")
        
        # Write dataframe to sheet
        for r in dataframe_to_rows(bank_df, index=False, header=True):
            ws.append(r)
        
        # Format header
        self._format_header_row(ws)
        
        # Auto-size columns
        self._auto_size_columns(ws)

    def add_monthly_sheet(self, monthly_df: pd.DataFrame) -> None:
        """
        Add monthly summary sheet
        
        Args:
            monthly_df: Monthly summary DataFrame
        """
        ws = self.workbook.create_sheet("Monthly Summary")
        
        # Write dataframe to sheet
        for r in dataframe_to_rows(monthly_df, index=False, header=True):
            ws.append(r)
        
        # Format header
        self._format_header_row(ws)
        
        # Format currency columns
        currency_cols = ['B', 'C', 'D']  # Total_Invoiced, Original_Deduction, Recalculated_Deduction
        for col in currency_cols:
            for cell in ws[col]:
                if cell.row > 1:  # Skip header
                    cell.number_format = self.config.currency_format
        
        # Auto-size columns
        self._auto_size_columns(ws)

    def save_workbook(self, output_path: Path) -> None:
        """
        Save workbook to file
        
        Args:
            output_path: Path to save workbook
        """
        if self.workbook is None:
            raise ValueError("Workbook not created. Call create_workbook() first.")
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)
        print(f"✓ Report saved: {output_path}")

    def _format_header_row(self, worksheet) -> None:
        """
        Format header row with consistent styling
        
        Args:
            worksheet: OpenPyXL worksheet
        """
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=self.config.excel_header_color,
                end_color=self.config.excel_header_color,
                fill_type="solid"
            )
            cell.font = Font(color=self.config.excel_header_font_color, bold=True)

    def _auto_size_columns(self, worksheet, columns: List[str] = None) -> None:
        """
        Auto-size columns based on content
        
        Args:
            worksheet: OpenPyXL worksheet
            columns: List of column letters to auto-size (default: all)
        """
        if columns is None:
            columns = [cell.column_letter for cell in worksheet[1]]
        
        for col in columns:
            max_length = 0
            for cell in worksheet[col]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col].width = adjusted_width

    def __repr__(self) -> str:
        """String representation"""
        sheets = len(self.workbook.worksheets) if self.workbook else 0
        return f"ReportGenerator(sheets={sheets})"
